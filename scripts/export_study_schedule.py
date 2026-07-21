#!/usr/bin/env python3
"""Export study_schedule.json from L3_2027_Final_3.xlsx."""

from __future__ import annotations

import json
import re
import sys
from datetime import datetime, time
from pathlib import Path

try:
    import openpyxl
except ImportError as exc:  # pragma: no cover
    raise SystemExit("openpyxl is required: pip install openpyxl") from exc

ROOT = Path(__file__).resolve().parents[1]
DEFAULT_XLSX = Path.home() / "Desktop/CFA L3 Exam/Schedule/L3_2027_Final_3.xlsx"
OUT_PATH = ROOT / "CFAL3/Resources/study_schedule.json"


def kind_for(label: str) -> str:
    if label.startswith("D3:"):
        return "deep3"
    if label.startswith("MM Video:"):
        return "video"
    if label.startswith("MM Q:"):
        return "video"
    if label.startswith("CFAI Q:"):
        return "questions"
    if label.startswith("TAKE:"):
        return "mock"
    if label.startswith("REVIEW:"):
        return "mock"
    if label.startswith("Extra Review:"):
        return "review"
    return "study"


def book_for(label: str) -> int | None:
    match = re.search(r"\bB(\d+)-M", label)
    return int(match.group(1)) if match else None


def export(xlsx_path: Path) -> dict:
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb["2027"]

    time_cols: list[tuple[int, str]] = []
    for col in range(7, ws.max_column + 1):
        value = ws.cell(4, col).value
        if isinstance(value, time):
            time_cols.append((col, value.strftime("%H:%M")))

    days: list[dict] = []
    for row in range(5, ws.max_row + 1):
        raw_date = ws.cell(row, 2).value
        if not isinstance(raw_date, datetime):
            continue

        day_date = raw_date.date().isoformat()
        hours_val = ws.cell(row, 6).value
        hours = float(hours_val) if isinstance(hours_val, (int, float)) else 0.0

        note = None
        for col in (4, 5):
            cell = ws.cell(row, col).value
            if isinstance(cell, str) and cell.strip() and cell not in {"Holiday", "Happenings"}:
                if not re.match(r"^(D3|MM|CFAI|TAKE|REVIEW|Extra)", cell):
                    note = cell.strip()

        slots: list[tuple[str, str]] = []
        for col, start in time_cols:
            value = ws.cell(row, col).value
            if isinstance(value, str) and value.strip():
                slots.append((start, value.strip()))

        blocks: list[dict] = []
        index = 0
        while index < len(slots):
            start, label = slots[index]
            end = index + 1
            while end < len(slots) and slots[end][1] == label:
                end += 1
            block = {
                "start": start,
                "label": label,
                "minutes": (end - index) * 15,
                "kind": kind_for(label),
            }
            book = book_for(label)
            if book is not None:
                block["book"] = book
            blocks.append(block)
            index = end

        days.append({"date": day_date, "hours": hours, "note": note, "blocks": blocks})

    wb.close()
    return {
        "version": 1,
        "source": xlsx_path.name,
        "exam_date": "2027-02-20",
        "total_planned_hours": 444,
        "days": days,
    }


def main(argv: list[str]) -> int:
    xlsx = Path(argv[1]) if len(argv) > 1 else DEFAULT_XLSX
    if not xlsx.exists():
        print(f"Workbook not found: {xlsx}", file=sys.stderr)
        return 1

    payload = export(xlsx)
    total = sum(day["hours"] for day in payload["days"])
    if len(payload["days"]) != 230 or abs(total - 444.0) > 0.001:
        print(
            f"Unexpected export: {len(payload['days'])} days, {total} hours",
            file=sys.stderr,
        )
        return 1

    OUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    OUT_PATH.write_text(json.dumps(payload, indent=2) + "\n", encoding="utf-8")
    print(f"Wrote {OUT_PATH} ({len(payload['days'])} days, {total} hours)")
    return 0


if __name__ == "__main__":
    raise SystemExit(main(sys.argv))
